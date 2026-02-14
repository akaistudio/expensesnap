"""
ExpenseSnap - Multi-Company Receipt Scanner (v4 - PostgreSQL)
=============================================================
One URL, multiple companies, isolated data.
PostgreSQL for reliable, backed-up data storage.

Roles:
  - Super Admin (first user): Creates companies, sees everything
  - Company Admin: Manages their team, sees their company data
  - Member: Uploads/views their company data
"""

import os, json, base64, uuid, hashlib, secrets
from datetime import datetime, timedelta
from pathlib import Path
from io import BytesIO
from functools import wraps
from urllib.request import urlopen

import anthropic
import fitz
import psycopg2
from psycopg2.extras import RealDictCursor
try:
    import pillow_heif
    pillow_heif.register_heif_opener()
except ImportError:
    pass  # HEIC support optional
from flask import Flask, request, jsonify, send_file, render_template_string, session, redirect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://localhost/expensesnap')
UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
MODEL = "claude-sonnet-4-5-20250929"

# ── Database ───────────────────────────────────────────────────
def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def init_db():
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS companies (
        id VARCHAR(36) PRIMARY KEY, name VARCHAR(255) NOT NULL,
        home_currency VARCHAR(10) DEFAULT 'USD',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS users (
        id VARCHAR(36) PRIMARY KEY, name VARCHAR(255) NOT NULL,
        email VARCHAR(255) UNIQUE NOT NULL, password_hash VARCHAR(255) NOT NULL,
        role VARCHAR(50) DEFAULT 'member', company_id VARCHAR(36),
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS expenses (
        id VARCHAR(36) PRIMARY KEY, date VARCHAR(20), vendor VARCHAR(255),
        location VARCHAR(255), category VARCHAR(100),
        subtotal DOUBLE PRECISION DEFAULT 0, tax DOUBLE PRECISION DEFAULT 0,
        tip DOUBLE PRECISION DEFAULT 0, total DOUBLE PRECISION DEFAULT 0,
        total_home DOUBLE PRECISION DEFAULT 0, total_usd DOUBLE PRECISION DEFAULT 0,
        payment_method VARCHAR(100), currency VARCHAR(10) DEFAULT 'USD',
        items TEXT, uploaded_by VARCHAR(255) DEFAULT 'default',
        company_id VARCHAR(36), receipt_image VARCHAR(255),
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS invite_codes (
        code VARCHAR(100) PRIMARY KEY, company_id VARCHAR(36),
        role VARCHAR(50) DEFAULT 'member', created_by VARCHAR(36),
        used_by VARCHAR(36), used_at TIMESTAMP,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    # Add columns if upgrading from older version
    for col, tbl, default in [
        ('home_currency', 'companies', "'USD'"),
        ('total_home', 'expenses', '0'),
        ('total_usd', 'expenses', '0'),
    ]:
        try:
            cur.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {'VARCHAR(10)' if 'currency' in col else 'DOUBLE PRECISION'} DEFAULT {default}")
        except Exception:
            conn.rollback()
    conn.commit(); cur.close(); conn.close()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# ── Currency Conversion ────────────────────────────────────────
_rate_cache = {}
_rate_cache_time = None

def get_exchange_rates():
    """Fetch exchange rates from free API, cache for 1 hour"""
    global _rate_cache, _rate_cache_time
    if _rate_cache and _rate_cache_time and (datetime.now() - _rate_cache_time).seconds < 3600:
        return _rate_cache
    try:
        url = "https://api.exchangerate-data.com/latest?base=USD"
        # Try free API
        try:
            response = urlopen("https://open.er-api.com/v6/latest/USD", timeout=5)
            data = json.loads(response.read())
            _rate_cache = data.get('rates', {})
        except Exception:
            # Fallback rates if API is down
            _rate_cache = {'USD':1,'CAD':1.36,'EUR':0.92,'GBP':0.79,'INR':83.5,'AUD':1.53,'JPY':149.5,'CHF':0.88,'SGD':1.34,'AED':3.67}
        _rate_cache_time = datetime.now()
    except Exception:
        if not _rate_cache:
            _rate_cache = {'USD':1,'CAD':1.36,'EUR':0.92,'GBP':0.79,'INR':83.5,'AUD':1.53,'JPY':149.5,'CHF':0.88,'SGD':1.34,'AED':3.67}
    return _rate_cache

def convert_currency(amount, from_curr, to_curr):
    """Convert amount between currencies"""
    if from_curr == to_curr or amount == 0:
        return round(amount, 2)
    rates = get_exchange_rates()
    from_curr = from_curr.upper()
    to_curr = to_curr.upper()
    # Convert to USD first, then to target
    from_rate = rates.get(from_curr, 1)
    to_rate = rates.get(to_curr, 1)
    usd_amount = amount / from_rate
    return round(usd_amount * to_rate, 2)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect('/login') if not request.path.startswith('/api/') else (jsonify({"error": "Not logged in"}), 401)
        return f(*args, **kwargs)
    return decorated

def is_super_admin(): return session.get('user_role') == 'super_admin'
def is_company_admin(): return session.get('user_role') in ('super_admin', 'company_admin')

# ── Claude API ─────────────────────────────────────────────────
def extract_receipt(image_list, media_type="image/jpeg"):
    client = anthropic.Anthropic()
    prompt = """Analyze this receipt/invoice (may be multiple pages) and extract ALL information.
Look across ALL pages carefully.
Return ONLY a valid JSON object with these exact keys:
{
  "date": "YYYY-MM-DD format, or empty string if not found",
  "vendor": "Business/restaurant name",
  "location": "City, State/Province or City, Country",
  "category": "One of: Food & Dining, Groceries, Air Travel, Cab & Rideshare, Hotel & Accommodation, Shopping & Retail, Utilities, Entertainment, Office & Business, Healthcare, Fuel & Parking, Other",
  "subtotal": 0.00, "tax": 0.00, "tip": 0.00, "total": 0.00,
  "payment_method": "e.g. Visa ****1234, Cash, etc.",
  "currency": "3-letter code e.g. CAD, USD, EUR, INR, GBP",
  "items": "List EVERY line item with its price. Format: 'Item Name (price), Item Name (price), ...'. Include quantities if shown. Example: '2x Cappuccino (6.00), Caesar Salad (12.50), Garlic Bread (5.00)'. Do NOT just show the total — list each item separately."
}
IMPORTANT:
- List ALL individual items in the items field, not just the total
- For tax: include the full tax amount. If multiple taxes (VAT, GST, CGST, SGST, service charge), add them all together
- Use 0.00 for missing amounts
- Return ONLY JSON, no other text."""
    content = []
    if isinstance(image_list, list):
        for img_bytes, mt in image_list:
            image_data = base64.standard_b64encode(img_bytes).decode("utf-8")
            content.append({"type": "image", "source": {"type": "base64", "media_type": mt, "data": image_data}})
    else:
        image_data = base64.standard_b64encode(image_list).decode("utf-8")
        content.append({"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_data}})
    content.append({"type": "text", "text": prompt})
    response = client.messages.create(model=MODEL, max_tokens=1000, messages=[{"role": "user", "content": content}])
    text = response.content[0].text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        text = text.rsplit("```", 1)[0]
    return json.loads(text)

# ── Excel Export ───────────────────────────────────────────────
def generate_excel(expenses, company_name=""):
    wb = Workbook(); ws = wb.active; ws.title = "Expenses"
    headers = ["Date","Vendor","Location","Category","Subtotal","Tax","Tip","Total","Payment Method","Currency","Items","Uploaded By"]
    widths = [14,28,22,18,14,12,12,14,18,12,35,20]
    hfill = PatternFill('solid', fgColor='1F4E79')
    hfont = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    dfont = Font(name='Arial', size=10)
    border = Border(bottom=Side(style='thin', color='D9D9D9'))
    curr_fmt = '$#,##0.00'
    if company_name:
        ws.cell(row=1, column=1, value=company_name).font = Font(name='Arial', bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Exported {datetime.now().strftime('%Y-%m-%d')}").font = Font(name='Arial', size=10, color='888888')
        start_row = 4
    else:
        start_row = 1
    for i, (name, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=start_row, column=i, value=name)
        cell.font, cell.fill = hfont, hfill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[chr(64+i) if i < 27 else 'A'+chr(64+i-26)].width = w
    ws.row_dimensions[start_row].height = 28; ws.freeze_panes = f'A{start_row+1}'
    for r, exp in enumerate(expenses, start_row+1):
        vals = [exp['date'],exp['vendor'],exp['location'],exp['category'],exp['subtotal'],
                exp['tax'],exp['tip'],exp['total'],exp['payment_method'],exp['currency'],exp['items'],exp.get('uploaded_by','')]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font, cell.border = dfont, border
            if c in (5,6,7,8): cell.number_format = curr_fmt
            cell.alignment = Alignment(horizontal='left' if c in (11,12) else 'center')
    sr = start_row + len(expenses) + 2
    ws.cell(row=sr, column=7, value="TOTAL:").font = Font(name='Arial', bold=True, size=11)
    ws.cell(row=sr, column=8, value=f"=SUM(H{start_row+1}:H{start_row+len(expenses)})").font = Font(name='Arial', bold=True, size=11)
    ws.cell(row=sr, column=8).number_format = curr_fmt
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── Auth Routes ────────────────────────────────────────────────
@app.route('/login')
def login_page():
    if 'user_id' in session: return redirect('/')
    return render_template_string(LOGIN_HTML)

@app.route('/register')
def register_page():
    return render_template_string(REGISTER_HTML)

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json; name = data.get('name','').strip(); email = data.get('email','').strip().lower()
    password = data.get('password',''); invite_code = data.get('invite_code','').strip()
    if not all([name, email, password]): return jsonify({"error": "All fields are required"}), 400
    if len(password) < 6: return jsonify({"error": "Password must be at least 6 characters"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE email = %s", (email,))
    if cur.fetchone(): conn.close(); return jsonify({"error": "Email already registered"}), 400
    cur.execute("SELECT COUNT(*) as cnt FROM users"); user_count = cur.fetchone()['cnt']
    if user_count == 0:
        user_id = str(uuid.uuid4())
        cur.execute("INSERT INTO users (id,name,email,password_hash,role) VALUES (%s,%s,%s,%s,%s)",
                     (user_id, name, email, hash_password(password), 'super_admin'))
        conn.commit(); conn.close()
        session.update({'user_id': user_id, 'user_name': name, 'user_role': 'super_admin', 'company_id': None, 'company_name': 'All Companies'})
        return jsonify({"success": True, "message": "Welcome! You are the Super Admin.", "role": "super_admin"})
    else:
        if not invite_code: conn.close(); return jsonify({"error": "Invite code required. Ask your admin for one."}), 400
        cur.execute("SELECT * FROM invite_codes WHERE code = %s AND used_by IS NULL", (invite_code,))
        invite = cur.fetchone()
        if not invite: conn.close(); return jsonify({"error": "Invalid or already used invite code"}), 400
        user_id = str(uuid.uuid4()); role = invite['role']; company_id = invite['company_id']
        cur.execute("INSERT INTO users (id,name,email,password_hash,role,company_id) VALUES (%s,%s,%s,%s,%s,%s)",
                     (user_id, name, email, hash_password(password), role, company_id))
        cur.execute("UPDATE invite_codes SET used_by=%s, used_at=%s WHERE code=%s", (user_id, datetime.now().isoformat(), invite_code))
        cur.execute("SELECT name FROM companies WHERE id=%s", (company_id,)); company = cur.fetchone()
        conn.commit(); conn.close(); company_name = company['name'] if company else ''
        session.update({'user_id': user_id, 'user_name': name, 'user_role': role, 'company_id': company_id, 'company_name': company_name})
        return jsonify({"success": True, "message": f"Welcome to {company_name}!", "role": role})

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json; email = data.get('email','').strip().lower(); password = data.get('password','')
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE email=%s AND password_hash=%s", (email, hash_password(password)))
    user = cur.fetchone()
    if not user: conn.close(); return jsonify({"error": "Invalid email or password"}), 401
    company_name = ''
    if user['company_id']:
        cur.execute("SELECT name FROM companies WHERE id=%s", (user['company_id'],)); c = cur.fetchone()
        company_name = c['name'] if c else ''
    conn.close()
    session.update({'user_id': user['id'], 'user_name': user['name'], 'user_role': user['role'],
                    'company_id': user['company_id'], 'company_name': company_name or 'All Companies'})
    return jsonify({"success": True, "name": user['name'], "role": user['role'], "company": company_name})

@app.route('/api/logout', methods=['POST'])
def logout(): session.clear(); return jsonify({"success": True})

@app.route('/forgot-password')
def forgot_password_page():
    return render_template_string(FORGOT_PASSWORD_HTML)

@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    data = request.json; email = data.get('email','').strip().lower()
    if not email: return jsonify({"error": "Email is required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id, name FROM users WHERE email=%s", (email,))
    user = cur.fetchone()
    if not user: conn.close(); return jsonify({"error": "No account found with that email"}), 404
    token = secrets.token_urlsafe(32)
    expires = (datetime.now() + __import__('datetime').timedelta(hours=1)).isoformat()
    cur.execute("""CREATE TABLE IF NOT EXISTS password_resets (
        token VARCHAR(100) PRIMARY KEY, user_id VARCHAR(100), expires_at VARCHAR(50), used BOOLEAN DEFAULT FALSE)""")
    cur.execute("DELETE FROM password_resets WHERE user_id=%s", (user['id'],))
    cur.execute("INSERT INTO password_resets (token, user_id, expires_at) VALUES (%s, %s, %s)", (token, user['id'], expires))
    conn.commit(); conn.close()
    reset_url = f"/reset-password?token={token}"
    return jsonify({"success": True, "reset_url": reset_url, "name": user['name'], "expires": "1 hour"})

@app.route('/reset-password')
def reset_password_page():
    return render_template_string(RESET_PASSWORD_HTML)

@app.route('/api/reset-password', methods=['POST'])
def do_reset_password():
    data = request.json; token = data.get('token','').strip(); new_password = data.get('password','')
    if not token: return jsonify({"error": "Invalid reset link"}), 400
    if len(new_password) < 6: return jsonify({"error": "Password must be at least 6 characters"}), 400
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM password_resets WHERE token=%s AND used=FALSE", (token,))
    except:
        conn.close(); return jsonify({"error": "Invalid reset link"}), 400
    reset = cur.fetchone()
    if not reset: conn.close(); return jsonify({"error": "Invalid or expired reset link"}), 400
    from datetime import datetime as dt
    if dt.fromisoformat(reset['expires_at']) < dt.now():
        conn.close(); return jsonify({"error": "Reset link has expired. Please request a new one."}), 400
    cur.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hash_password(new_password), reset['user_id']))
    cur.execute("UPDATE password_resets SET used=TRUE WHERE token=%s", (token,))
    conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route('/api/me')
def get_me():
    if 'user_id' not in session: return jsonify({"logged_in": False}), 401
    return jsonify({"logged_in": True, "name": session.get('user_name'), "role": session.get('user_role'),
                    "company": session.get('company_name'), "company_id": session.get('company_id')})

# ── Company Management ─────────────────────────────────────────
@app.route('/api/companies', methods=['GET'])
@login_required
def list_companies():
    if not is_super_admin(): return jsonify({"error": "Super admin only"}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM companies ORDER BY created_at"); companies = cur.fetchall(); result = []
    for c in companies:
        cur.execute("SELECT COUNT(*) as cnt FROM users WHERE company_id=%s", (c['id'],)); uc = cur.fetchone()['cnt']
        cur.execute("SELECT COUNT(*) as cnt FROM expenses WHERE company_id=%s", (c['id'],)); ec = cur.fetchone()['cnt']
        cur.execute("SELECT COALESCE(SUM(total),0) as t FROM expenses WHERE company_id=%s", (c['id'],)); ts = float(cur.fetchone()['t'])
        result.append({**dict(c), 'user_count': uc, 'expense_count': ec, 'total_spent': ts})
    conn.close(); return jsonify(result)

@app.route('/api/companies', methods=['POST'])
@login_required
def create_company():
    if not is_super_admin(): return jsonify({"error": "Super admin only"}), 403
    name = request.json.get('name','').strip()
    home_currency = request.json.get('home_currency','USD').strip().upper()
    if not name: return jsonify({"error": "Company name required"}), 400
    company_id = str(uuid.uuid4())[:8]; code = secrets.token_urlsafe(8)
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO companies (id,name,home_currency) VALUES (%s,%s,%s)", (company_id, name, home_currency))
    cur.execute("INSERT INTO invite_codes (code,company_id,role,created_by) VALUES (%s,%s,%s,%s)", (code, company_id, 'company_admin', session['user_id']))
    conn.commit(); conn.close()
    return jsonify({"success": True, "company_id": company_id, "admin_invite_code": code, "message": f"Company '{name}' created! Currency: {home_currency}. Admin invite: {code}"})

@app.route('/api/companies/<company_id>', methods=['DELETE'])
@login_required
def delete_company(company_id):
    if not is_super_admin(): return jsonify({"error": "Super admin only"}), 403
    conn = get_db(); cur = conn.cursor()
    for t in ['expenses','users','invite_codes']: cur.execute(f"DELETE FROM {t} WHERE company_id=%s", (company_id,))
    cur.execute("DELETE FROM companies WHERE id=%s", (company_id,)); conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route('/api/companies/<company_id>', methods=['PUT'])
@login_required
def edit_company(company_id):
    """Edit company settings - accessible by super admin or that company's admin"""
    if not is_company_admin(): return jsonify({"error": "Admin access required"}), 403
    if not is_super_admin() and session.get('company_id') != company_id:
        return jsonify({"error": "Can only edit your own company"}), 403
    data = request.json or {}
    conn = get_db(); cur = conn.cursor()
    fields, values = [], []
    if 'name' in data and data['name'].strip():
        fields.append("name=%s"); values.append(data['name'].strip())
    if 'home_currency' in data and data['home_currency'].strip():
        fields.append("home_currency=%s"); values.append(data['home_currency'].strip().upper())
    if not fields: conn.close(); return jsonify({"error": "Nothing to update"}), 400
    values.append(company_id)
    cur.execute(f"UPDATE companies SET {','.join(fields)} WHERE id=%s", values)
    conn.commit(); conn.close()
    # Update session if editing own company
    if session.get('company_id') == company_id:
        if 'name' in data: session['company_name'] = data['name'].strip()
    return jsonify({"success": True})

@app.route('/api/companies/<company_id>/recalculate', methods=['POST'])
@login_required
def recalculate_expenses(company_id):
    """Recalculate all converted amounts for a company using current exchange rates"""
    if not is_company_admin(): return jsonify({"error": "Admin access required"}), 403
    if not is_super_admin() and session.get('company_id') != company_id:
        return jsonify({"error": "Can only recalculate your own company"}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT home_currency FROM companies WHERE id=%s", (company_id,))
    comp = cur.fetchone()
    if not comp: conn.close(); return jsonify({"error": "Company not found"}), 404
    home_currency = comp.get('home_currency', 'USD') or 'USD'
    cur.execute("SELECT id, total, currency FROM expenses WHERE company_id=%s", (company_id,))
    expenses = cur.fetchall(); updated = 0
    for e in expenses:
        bill_curr = (e.get('currency') or 'USD').upper()
        total = float(e.get('total') or 0)
        total_home = convert_currency(total, bill_curr, home_currency)
        total_usd = convert_currency(total, bill_curr, 'USD')
        cur.execute("UPDATE expenses SET total_home=%s, total_usd=%s WHERE id=%s", (total_home, total_usd, e['id']))
        updated += 1
    conn.commit(); conn.close()
    return jsonify({"success": True, "updated": updated, "home_currency": home_currency})

@app.route('/api/my-company')
@login_required
def get_my_company():
    """Get current user's company settings"""
    company_id = session.get('company_id')
    if not company_id: return jsonify({"error": "No company"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM companies WHERE id=%s", (company_id,))
    comp = cur.fetchone(); conn.close()
    if not comp: return jsonify({"error": "Company not found"}), 404
    return jsonify(dict(comp))

# ── Invite Codes ───────────────────────────────────────────────
@app.route('/api/invite', methods=['POST'])
@login_required
def create_invite():
    if not is_company_admin(): return jsonify({"error": "Admin access required"}), 403
    role = request.json.get('role', 'member') if request.json else 'member'
    if is_super_admin():
        company_id = request.json.get('company_id') if request.json else None
        if not company_id: return jsonify({"error": "Select a company"}), 400
        if role not in ('company_admin','member'): role = 'member'
    else: company_id = session.get('company_id'); role = 'member'
    code = secrets.token_urlsafe(8); conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO invite_codes (code,company_id,role,created_by) VALUES (%s,%s,%s,%s)", (code, company_id, role, session['user_id']))
    conn.commit(); conn.close(); return jsonify({"success": True, "code": code})

# ── Team ───────────────────────────────────────────────────────
@app.route('/api/team')
@login_required
def get_team():
    if not is_company_admin(): return jsonify({"error": "Admin access required"}), 403
    conn = get_db(); cur = conn.cursor()
    if is_super_admin():
        cid = request.args.get('company_id')
        if cid:
            cur.execute("SELECT id,name,email,role,created_at FROM users WHERE company_id=%s ORDER BY created_at", (cid,)); users = cur.fetchall()
            cur.execute("SELECT code,role,created_at FROM invite_codes WHERE company_id=%s AND used_by IS NULL", (cid,)); invites = cur.fetchall()
        else:
            cur.execute("SELECT id,name,email,role,company_id,created_at FROM users ORDER BY created_at"); users = cur.fetchall()
            cur.execute("SELECT code,company_id,role,created_at FROM invite_codes WHERE used_by IS NULL"); invites = cur.fetchall()
    else:
        cid = session.get('company_id')
        cur.execute("SELECT id,name,email,role,created_at FROM users WHERE company_id=%s ORDER BY created_at", (cid,)); users = cur.fetchall()
        cur.execute("SELECT code,role,created_at FROM invite_codes WHERE company_id=%s AND used_by IS NULL", (cid,)); invites = cur.fetchall()
    conn.close(); return jsonify({"users": [dict(u) for u in users], "pending_invites": [dict(i) for i in invites]})

@app.route('/api/team/<user_id>', methods=['DELETE'])
@login_required
def remove_member(user_id):
    if not is_company_admin(): return jsonify({"error": "Admin access required"}), 403
    if user_id == session['user_id']: return jsonify({"error": "Cannot remove yourself"}), 400
    conn = get_db(); cur = conn.cursor()
    if not is_super_admin():
        cur.execute("SELECT company_id FROM users WHERE id=%s", (user_id,)); u = cur.fetchone()
        if not u or u['company_id'] != session.get('company_id'): conn.close(); return jsonify({"error": "Access denied"}), 403
    cur.execute("DELETE FROM users WHERE id=%s", (user_id,)); conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route('/api/team/<user_id>/reset-password', methods=['POST'])
@login_required
def reset_password(user_id):
    if not is_company_admin(): return jsonify({"error": "Admin access required"}), 403
    new_password = request.json.get('password','').strip() if request.json else ''
    if len(new_password) < 6: return jsonify({"error": "Password must be at least 6 characters"}), 400
    conn = get_db(); cur = conn.cursor()
    if not is_super_admin():
        cur.execute("SELECT company_id FROM users WHERE id=%s", (user_id,)); u = cur.fetchone()
        if not u or u['company_id'] != session.get('company_id'): conn.close(); return jsonify({"error": "Access denied"}), 403
    cur.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hash_password(new_password), user_id))
    conn.commit(); conn.close()
    return jsonify({"success": True})
@app.route('/api/upload', methods=['POST'])
@login_required
def upload_receipt():
    if 'receipt' not in request.files: return jsonify({"error": "No file uploaded"}), 400
    file = request.files['receipt']
    ext_map = {'.jpg':'image/jpeg','.jpeg':'image/jpeg','.png':'image/png','.webp':'image/webp','.gif':'image/gif','.heic':'image/heic','.heif':'image/heic','.pdf':'application/pdf'}
    ext = Path(file.filename).suffix.lower(); media_type = ext_map.get(ext, 'image/jpeg'); image_bytes = file.read()

    # Convert HEIC to JPEG using Pillow
    if ext in ('.heic', '.heif'):
        try:
            from PIL import Image
            img = Image.open(BytesIO(image_bytes))
            buf = BytesIO()
            img.convert('RGB').save(buf, format='JPEG', quality=85)
            image_bytes = buf.getvalue()
            media_type = 'image/jpeg'
            ext = '.jpg'
        except Exception as e:
            return jsonify({"error": f"Failed to convert HEIC: {str(e)}"}), 400

    # Compress large images (over 1.5MB) to reduce size
    if ext not in ('.pdf',) and len(image_bytes) > 1.5 * 1024 * 1024:
        try:
            from PIL import Image
            img = Image.open(BytesIO(image_bytes))
            # Resize if very large dimensions
            max_dim = 2000
            if max(img.size) > max_dim:
                img.thumbnail((max_dim, max_dim), Image.LANCZOS)
            buf = BytesIO()
            img.convert('RGB').save(buf, format='JPEG', quality=80)
            image_bytes = buf.getvalue()
            media_type = 'image/jpeg'
            ext = '.jpg'
        except Exception as e:
            pass  # If compression fails, try with original

    if ext == '.pdf':
        try:
            pdf_doc = fitz.open(stream=image_bytes, filetype="pdf"); page_images = []
            for i in range(min(len(pdf_doc), 10)):
                pix = pdf_doc[i].get_pixmap(dpi=200); page_images.append((pix.tobytes("png"), "image/png"))
            pdf_doc.close()
            img_id = str(uuid.uuid4())[:8]; img_path = UPLOAD_DIR / f"{img_id}.png"
            with open(img_path, 'wb') as f: f.write(page_images[0][0])
            try: data = extract_receipt(page_images)
            except Exception as e: return jsonify({"error": f"Failed to extract: {str(e)}"}), 500
        except Exception as e: return jsonify({"error": f"Failed to read PDF: {str(e)}"}), 400
    else:
        img_id = str(uuid.uuid4())[:8]; img_path = UPLOAD_DIR / f"{img_id}{ext}"
        with open(img_path, 'wb') as f: f.write(image_bytes)
        try: data = extract_receipt(image_bytes, media_type)
        except Exception as e: return jsonify({"error": f"Failed to extract: {str(e)}"}), 500
    expense_id = str(uuid.uuid4()); company_id = session.get('company_id'); uploader = session.get('user_name', 'unknown')
    conn = get_db(); cur = conn.cursor()
    # Get company home currency for conversion
    bill_currency = data.get('currency','USD').upper()
    home_currency = 'USD'
    if company_id:
        cur.execute("SELECT home_currency FROM companies WHERE id=%s", (company_id,))
        comp = cur.fetchone()
        if comp: home_currency = comp.get('home_currency','USD') or 'USD'
    total = float(data.get('total',0))
    total_home = convert_currency(total, bill_currency, home_currency)
    total_usd = convert_currency(total, bill_currency, 'USD')
    cur.execute("""INSERT INTO expenses (id,date,vendor,location,category,subtotal,tax,tip,total,total_home,total_usd,payment_method,currency,items,uploaded_by,company_id,receipt_image)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                 (expense_id, data.get('date',''), data.get('vendor',''), data.get('location',''),
                  data.get('category',''), data.get('subtotal',0), data.get('tax',0), data.get('tip',0),
                  total, total_home, total_usd, data.get('payment_method',''), bill_currency,
                  data.get('items',''), uploader, company_id, f"{img_id}{'.png' if ext=='.pdf' else ext}"))
    conn.commit(); conn.close()
    data['id'] = expense_id; data['uploaded_by'] = uploader
    data['total_home'] = total_home; data['total_usd'] = total_usd; data['home_currency'] = home_currency
    return jsonify({"success": True, "expense": data})

@app.route('/api/expenses')
@login_required
def get_expenses():
    conn = get_db(); cur = conn.cursor()
    if is_super_admin():
        cid = request.args.get('company_id')
        if cid: cur.execute("SELECT e.*, c.name as company_name FROM expenses e LEFT JOIN companies c ON e.company_id=c.id WHERE e.company_id=%s ORDER BY e.date DESC", (cid,))
        else: cur.execute("SELECT e.*, c.name as company_name FROM expenses e LEFT JOIN companies c ON e.company_id=c.id ORDER BY e.date DESC")
    else: cur.execute("SELECT * FROM expenses WHERE company_id=%s ORDER BY date DESC", (session.get('company_id'),))
    rows = cur.fetchall(); conn.close(); return jsonify([dict(r) for r in rows])

@app.route('/api/expenses/<expense_id>', methods=['DELETE'])
@login_required
def delete_expense(expense_id):
    conn = get_db(); cur = conn.cursor()
    if not is_super_admin():
        cur.execute("SELECT company_id FROM expenses WHERE id=%s", (expense_id,)); exp = cur.fetchone()
        if not exp or exp['company_id'] != session.get('company_id'): conn.close(); return jsonify({"error": "Access denied"}), 403
    cur.execute("DELETE FROM expenses WHERE id=%s", (expense_id,)); conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route('/api/expenses/<expense_id>', methods=['PUT'])
@login_required
def update_expense(expense_id):
    data = request.json; conn = get_db(); cur = conn.cursor(); fields, values = [], []
    for key in ['date','vendor','location','category','subtotal','tax','tip','total','payment_method','currency','items']:
        if key in data: fields.append(f"{key}=%s"); values.append(data[key])
    values.append(expense_id)
    cur.execute(f"UPDATE expenses SET {','.join(fields)} WHERE id=%s", values); conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route('/api/dashboard')
@login_required
def dashboard_data():
    conn = get_db(); cur = conn.cursor()
    if is_super_admin():
        cid = request.args.get('company_id')
        if cid: cur.execute("SELECT * FROM expenses WHERE company_id=%s ORDER BY date DESC", (cid,))
        else: cur.execute("SELECT * FROM expenses ORDER BY date DESC")
    else: cur.execute("SELECT * FROM expenses WHERE company_id=%s ORDER BY date DESC", (session.get('company_id'),))
    expenses = [dict(r) for r in cur.fetchall()]; conn.close()
    total_bill = sum(float(e['total'] or 0) for e in expenses)
    total_home = sum(float(e.get('total_home') or e['total'] or 0) for e in expenses)
    total_usd = sum(float(e.get('total_usd') or e['total'] or 0) for e in expenses)
    # Get home currency
    home_currency = 'USD'
    if expenses and expenses[0].get('company_id'):
        try:
            c2 = get_db(); cur2 = c2.cursor()
            cur2.execute("SELECT home_currency FROM companies WHERE id=%s", (expenses[0]['company_id'],))
            comp = cur2.fetchone(); c2.close()
            if comp: home_currency = comp.get('home_currency','USD') or 'USD'
        except: pass
    by_category, by_month, by_user = {}, {}, {}
    for e in expenses:
        cat = e['category'] or 'Other'; by_category[cat] = by_category.get(cat, 0) + float(e['total'] or 0)
        month = e['date'][:7] if e['date'] else 'Unknown'; by_month[month] = by_month.get(month, 0) + float(e['total'] or 0)
        user = e.get('uploaded_by', 'unknown'); by_user[user] = by_user.get(user, 0) + float(e['total'] or 0)
    return jsonify({"total": total_bill, "total_home": total_home, "total_usd": total_usd,
                    "home_currency": home_currency, "count": len(expenses), "by_category": by_category,
                    "by_month": dict(sorted(by_month.items())), "by_user": by_user, "recent": expenses[:10]})

@app.route('/api/export')
@login_required
def export_excel():
    conn = get_db(); cur = conn.cursor(); company_name = session.get('company_name', '')
    if is_super_admin():
        cid = request.args.get('company_id')
        if cid:
            cur.execute("SELECT * FROM expenses WHERE company_id=%s ORDER BY date ASC", (cid,)); rows = cur.fetchall()
            cur.execute("SELECT name FROM companies WHERE id=%s", (cid,)); c = cur.fetchone(); company_name = c['name'] if c else ''
        else: cur.execute("SELECT * FROM expenses ORDER BY date ASC"); rows = cur.fetchall(); company_name = 'All Companies'
    else: cur.execute("SELECT * FROM expenses WHERE company_id=%s ORDER BY date ASC", (session.get('company_id'),)); rows = cur.fetchall()
    conn.close(); expenses = [dict(r) for r in rows]
    if not expenses: return jsonify({"error": "No expenses to export"}), 400
    buf = generate_excel(expenses, company_name); today = datetime.now().strftime('%Y-%m-%d')
    return send_file(buf, download_name=f"expenses_{today}.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/')
@login_required
def index():
    return render_template_string(MAIN_HTML, user_name=session.get('user_name',''),
                                  user_role=session.get('user_role','member'), company_name=session.get('company_name',''),
                                  company_id=session.get('company_id',''))


# ── Login HTML ─────────────────────────────────────────────────
LOGIN_HTML = r"""
<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no">
<title>ExpenseSnap - Login</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
:root{--bg:#0B0F1A;--surface:#141926;--border:#2A3148;--text:#E8ECF4;--text2:#8B95B0;
--accent:#6C5CE7;--accent2:#A29BFE;--green:#00D2A0;--red:#FF6B6B;--radius:16px}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;
display:flex;align-items:center;justify-content:center}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);
padding:48px;width:100%;max-width:420px;margin:20px}
.logo{font-size:28px;font-weight:700;text-align:center;margin-bottom:8px;
background:linear-gradient(135deg,var(--accent),var(--green));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.subtitle{text-align:center;color:var(--text2);font-size:14px;margin-bottom:32px}
label{font-size:13px;color:var(--text2);display:block;margin-bottom:6px;margin-top:16px}
input{width:100%;padding:12px 16px;background:var(--bg);border:1px solid var(--border);
border-radius:10px;color:var(--text);font-family:inherit;font-size:14px;outline:none}
input:focus{border-color:var(--accent)}
.btn{width:100%;padding:14px;background:var(--accent);color:white;border:none;border-radius:10px;
font-family:inherit;font-size:15px;font-weight:600;cursor:pointer;margin-top:24px;transition:all 0.2s}
.btn:hover{background:#5A4BD1}
.switch{text-align:center;margin-top:20px;font-size:13px;color:var(--text2)}
.switch a{color:var(--accent2);text-decoration:none}
.error{background:rgba(255,107,107,0.1);color:var(--red);padding:12px;border-radius:10px;font-size:13px;margin-top:16px;display:none}
</style></head><body>
<div class="card">
<div class="logo">ExpenseSnap</div>
<div class="subtitle">Sign in to your account</div>
<div class="error" id="error"></div>
<form onsubmit="handleLogin(event)">
<label>Email</label><input type="email" id="email" required placeholder="you@company.com">
<label>Password</label><input type="password" id="password" required placeholder="Enter password">
<button type="submit" class="btn">Sign In</button>
</form>
<div class="switch">New here? <a href="/register">Create an account</a></div>
<div class="switch" style="margin-top:10px"><a href="/forgot-password">Forgot password?</a></div>
</div>
<script>
async function handleLogin(e){e.preventDefault();const err=document.getElementById('error');err.style.display='none';
try{const res=await fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},
body:JSON.stringify({email:document.getElementById('email').value,password:document.getElementById('password').value})});
const data=await res.json();if(data.success){window.location.href='/'}
else{err.textContent=data.error;err.style.display='block'}}
catch(e){err.textContent='Connection error';err.style.display='block'}}
</script></body></html>"""

# ── Register HTML ──────────────────────────────────────────────
REGISTER_HTML = r"""
<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no">
<title>ExpenseSnap - Register</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
:root{--bg:#0B0F1A;--surface:#141926;--border:#2A3148;--text:#E8ECF4;--text2:#8B95B0;
--accent:#6C5CE7;--accent2:#A29BFE;--green:#00D2A0;--red:#FF6B6B;--radius:16px}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;
display:flex;align-items:center;justify-content:center}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);
padding:48px;width:100%;max-width:420px;margin:20px}
.logo{font-size:28px;font-weight:700;text-align:center;margin-bottom:8px;
background:linear-gradient(135deg,var(--accent),var(--green));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.subtitle{text-align:center;color:var(--text2);font-size:14px;margin-bottom:32px}
.note{background:rgba(108,92,231,0.1);color:var(--accent2);padding:12px;border-radius:10px;font-size:12px;margin-bottom:16px;line-height:1.5}
label{font-size:13px;color:var(--text2);display:block;margin-bottom:6px;margin-top:16px}
input{width:100%;padding:12px 16px;background:var(--bg);border:1px solid var(--border);
border-radius:10px;color:var(--text);font-family:inherit;font-size:14px;outline:none}
input:focus{border-color:var(--accent)}
.btn{width:100%;padding:14px;background:var(--accent);color:white;border:none;border-radius:10px;
font-family:inherit;font-size:15px;font-weight:600;cursor:pointer;margin-top:24px;transition:all 0.2s}
.btn:hover{background:#5A4BD1}
.switch{text-align:center;margin-top:20px;font-size:13px;color:var(--text2)}
.switch a{color:var(--accent2);text-decoration:none}
.error{background:rgba(255,107,107,0.1);color:var(--red);padding:12px;border-radius:10px;font-size:13px;margin-top:16px;display:none}
</style></head><body>
<div class="card">
<div class="logo">ExpenseSnap</div>
<div class="subtitle">Create your account</div>
<div class="note">🔑 First person to register becomes the <strong>Super Admin</strong> (no invite needed).<br>
Everyone else needs an invite code from their company admin.</div>
<div class="error" id="error"></div>
<form onsubmit="handleRegister(event)">
<label>Full Name</label><input type="text" id="name" required placeholder="Your name">
<label>Email</label><input type="email" id="email" required placeholder="you@company.com">
<label>Password</label><input type="password" id="password" required placeholder="Min 6 characters" minlength="6">
<label>Invite Code <span style="color:var(--text2)">(not needed for first user)</span></label>
<input type="text" id="invite_code" placeholder="Ask your admin for this">
<button type="submit" class="btn">Create Account</button>
</form>
<div class="switch">Already have an account? <a href="/login">Sign in</a></div>
</div>
<script>
async function handleRegister(e){e.preventDefault();const err=document.getElementById('error');err.style.display='none';
try{const res=await fetch('/api/register',{method:'POST',headers:{'Content-Type':'application/json'},
body:JSON.stringify({name:document.getElementById('name').value,email:document.getElementById('email').value,
password:document.getElementById('password').value,invite_code:document.getElementById('invite_code').value})});
const data=await res.json();if(data.success){window.location.href='/'}
else{err.textContent=data.error;err.style.display='block'}}
catch(e){err.textContent='Connection error';err.style.display='block'}}
</script></body></html>"""

# ── Forgot Password HTML ───────────────────────────────────────
FORGOT_PASSWORD_HTML = r"""
<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no">
<title>ExpenseSnap - Forgot Password</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
:root{--bg:#0B0F1A;--surface:#141926;--border:#2A3148;--text:#E8ECF4;--text2:#8B95B0;
--accent:#6C5CE7;--accent2:#A29BFE;--green:#00D2A0;--red:#FF6B6B;--radius:16px}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;
display:flex;align-items:center;justify-content:center}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);
padding:48px;width:100%;max-width:420px;margin:20px}
.logo{font-size:28px;font-weight:700;text-align:center;margin-bottom:8px;
background:linear-gradient(135deg,var(--accent),var(--green));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.subtitle{text-align:center;color:var(--text2);font-size:14px;margin-bottom:32px}
label{font-size:13px;color:var(--text2);display:block;margin-bottom:6px;margin-top:16px}
input{width:100%;padding:12px 16px;background:var(--bg);border:1px solid var(--border);
border-radius:10px;color:var(--text);font-family:inherit;font-size:14px;outline:none}
input:focus{border-color:var(--accent)}
.btn{width:100%;padding:14px;background:var(--accent);color:white;border:none;border-radius:10px;
font-family:inherit;font-size:15px;font-weight:600;cursor:pointer;margin-top:24px;transition:all 0.2s}
.btn:hover{background:#5A4BD1}
.switch{text-align:center;margin-top:20px;font-size:13px;color:var(--text2)}
.switch a{color:var(--accent2);text-decoration:none}
.error{background:rgba(255,107,107,0.1);color:var(--red);padding:12px;border-radius:10px;font-size:13px;margin-top:16px;display:none}
.success-box{background:rgba(0,210,160,0.1);border:1px solid rgba(0,210,160,0.3);color:var(--green);
padding:16px;border-radius:10px;font-size:13px;margin-top:16px;display:none;line-height:1.6}
.reset-link{background:var(--bg);padding:10px;border-radius:8px;margin-top:10px;word-break:break-all;
font-family:monospace;font-size:12px;color:var(--accent2);cursor:pointer;border:1px solid var(--border)}
.reset-link:hover{border-color:var(--accent)}
</style></head><body>
<div class="card">
<div class="logo">ExpenseSnap</div>
<div class="subtitle">Reset your password</div>
<div class="error" id="error"></div>
<div class="success-box" id="success">
<div id="successMsg"></div>
<div class="reset-link" id="resetLink" onclick="copyLink()" title="Click to copy"></div>
<div style="text-align:center;margin-top:8px;font-size:11px;color:var(--text2)">Click the link above to copy it, then open it in your browser</div>
</div>
<form id="forgotForm" onsubmit="handleForgot(event)">
<label>Email Address</label><input type="email" id="email" required placeholder="Enter your registered email">
<button type="submit" class="btn" id="submitBtn">Send Reset Link</button>
</form>
<div class="switch"><a href="/login">Back to Sign In</a></div>
</div>
<script>
async function handleForgot(e){e.preventDefault();
const err=document.getElementById('error');const success=document.getElementById('success');
err.style.display='none';success.style.display='none';
const btn=document.getElementById('submitBtn');btn.textContent='Sending...';btn.disabled=true;
try{const res=await fetch('/api/forgot-password',{method:'POST',headers:{'Content-Type':'application/json'},
body:JSON.stringify({email:document.getElementById('email').value})});
const data=await res.json();
if(data.success){
document.getElementById('forgotForm').style.display='none';
document.getElementById('successMsg').innerHTML=`Hi <strong>${data.name}</strong>! Your reset link is ready (expires in ${data.expires}):`;
const fullUrl=window.location.origin+data.reset_url;
document.getElementById('resetLink').textContent=fullUrl;
document.getElementById('resetLink').dataset.url=fullUrl;
success.style.display='block';
}else{err.textContent=data.error;err.style.display='block'}}
catch(e){err.textContent='Connection error';err.style.display='block'}
btn.textContent='Send Reset Link';btn.disabled=false;}
function copyLink(){const url=document.getElementById('resetLink').dataset.url;
navigator.clipboard.writeText(url).then(()=>{document.getElementById('resetLink').style.borderColor='var(--green)';
setTimeout(()=>window.location.href=document.getElementById('resetLink').dataset.url,500)});}
</script></body></html>"""

# ── Reset Password HTML ────────────────────────────────────────
RESET_PASSWORD_HTML = r"""
<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no">
<title>ExpenseSnap - New Password</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
:root{--bg:#0B0F1A;--surface:#141926;--border:#2A3148;--text:#E8ECF4;--text2:#8B95B0;
--accent:#6C5CE7;--accent2:#A29BFE;--green:#00D2A0;--red:#FF6B6B;--radius:16px}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;
display:flex;align-items:center;justify-content:center}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);
padding:48px;width:100%;max-width:420px;margin:20px}
.logo{font-size:28px;font-weight:700;text-align:center;margin-bottom:8px;
background:linear-gradient(135deg,var(--accent),var(--green));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.subtitle{text-align:center;color:var(--text2);font-size:14px;margin-bottom:32px}
label{font-size:13px;color:var(--text2);display:block;margin-bottom:6px;margin-top:16px}
input{width:100%;padding:12px 16px;background:var(--bg);border:1px solid var(--border);
border-radius:10px;color:var(--text);font-family:inherit;font-size:14px;outline:none}
input:focus{border-color:var(--accent)}
.btn{width:100%;padding:14px;background:var(--accent);color:white;border:none;border-radius:10px;
font-family:inherit;font-size:15px;font-weight:600;cursor:pointer;margin-top:24px;transition:all 0.2s}
.btn:hover{background:#5A4BD1}
.switch{text-align:center;margin-top:20px;font-size:13px;color:var(--text2)}
.switch a{color:var(--accent2);text-decoration:none}
.error{background:rgba(255,107,107,0.1);color:var(--red);padding:12px;border-radius:10px;font-size:13px;margin-top:16px;display:none}
.success-box{background:rgba(0,210,160,0.1);border:1px solid rgba(0,210,160,0.3);color:var(--green);
padding:16px;border-radius:10px;font-size:14px;margin-top:16px;display:none;text-align:center}
</style></head><body>
<div class="card">
<div class="logo">ExpenseSnap</div>
<div class="subtitle">Set your new password</div>
<div class="error" id="error"></div>
<div class="success-box" id="success">Password updated! Redirecting to login...</div>
<form id="resetForm" onsubmit="handleReset(event)">
<label>New Password</label><input type="password" id="password" required placeholder="Minimum 6 characters" minlength="6">
<label>Confirm Password</label><input type="password" id="confirm" required placeholder="Re-enter password" minlength="6">
<button type="submit" class="btn" id="submitBtn">Update Password</button>
</form>
<div class="switch"><a href="/login">Back to Sign In</a></div>
</div>
<script>
const token=new URLSearchParams(window.location.search).get('token');
if(!token){document.getElementById('error').textContent='Invalid reset link. Please request a new one.';
document.getElementById('error').style.display='block';document.getElementById('resetForm').style.display='none';}
async function handleReset(e){e.preventDefault();
const err=document.getElementById('error');err.style.display='none';
const pwd=document.getElementById('password').value;const confirm=document.getElementById('confirm').value;
if(pwd!==confirm){err.textContent='Passwords do not match';err.style.display='block';return;}
const btn=document.getElementById('submitBtn');btn.textContent='Updating...';btn.disabled=true;
try{const res=await fetch('/api/reset-password',{method:'POST',headers:{'Content-Type':'application/json'},
body:JSON.stringify({token:token,password:pwd})});
const data=await res.json();
if(data.success){document.getElementById('resetForm').style.display='none';
document.getElementById('success').style.display='block';
setTimeout(()=>window.location.href='/login',2000);}
else{err.textContent=data.error;err.style.display='block';}}
catch(e){err.textContent='Connection error';err.style.display='block';}
btn.textContent='Update Password';btn.disabled=false;}
</script></body></html>"""

# ── Main App HTML ──────────────────────────────────────────────
MAIN_HTML = r"""
<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no">
<title>ExpenseSnap</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{--bg:#0B0F1A;--surface:#141926;--surface2:#1C2235;--border:#2A3148;--text:#E8ECF4;--text2:#8B95B0;
--accent:#6C5CE7;--accent2:#A29BFE;--green:#00D2A0;--red:#FF6B6B;--orange:#FDCB6E;--blue:#74B9FF;
--radius:16px;--radius-sm:10px}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;overflow-x:hidden}

.topbar{display:flex;align-items:center;justify-content:space-between;padding:20px 28px;
background:var(--surface);border-bottom:1px solid var(--border);position:sticky;top:0;z-index:100}
.logo{font-size:22px;font-weight:700;background:linear-gradient(135deg,var(--accent),var(--green));
-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.logo span{font-weight:400;opacity:0.7}
.topbar-right{display:flex;align-items:center;gap:12px}
.user-badge{background:var(--surface2);padding:8px 16px;border-radius:var(--radius-sm);font-size:13px;color:var(--text2)}
.user-badge strong{color:var(--text)}
.company-badge{background:rgba(0,210,160,0.1);color:var(--green);padding:6px 14px;border-radius:20px;font-size:12px;font-weight:600}
.btn{padding:10px 20px;border:none;border-radius:var(--radius-sm);font-family:inherit;font-size:14px;font-weight:600;cursor:pointer;transition:all 0.2s}
.btn-primary{background:var(--accent);color:white}.btn-primary:hover{background:#5A4BD1;transform:translateY(-1px)}
.btn-ghost{background:transparent;color:var(--text2);border:1px solid var(--border)}.btn-ghost:hover{color:var(--text);border-color:var(--text2)}
.btn-danger{background:transparent;color:var(--red);border:1px solid rgba(255,107,107,0.3)}.btn-danger:hover{background:rgba(255,107,107,0.1)}
.btn-sm{padding:6px 14px;font-size:12px}

.nav{display:flex;gap:0;background:var(--surface);border-bottom:1px solid var(--border);padding:0 28px}
.nav-tab{padding:16px 24px;font-size:14px;font-weight:500;color:var(--text2);cursor:pointer;border:none;
background:none;border-bottom:2px solid transparent;transition:all 0.2s;font-family:inherit}
.nav-tab:hover{color:var(--text)}.nav-tab.active{color:var(--accent2);border-bottom-color:var(--accent)}

.main{padding:28px;max-width:1200px;margin:0 auto}

/* Company selector for super admin */
.company-selector{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);
padding:16px 20px;margin-bottom:20px;display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.company-selector label{font-size:13px;color:var(--text2);font-weight:600}
.company-selector select{background:var(--bg);border:1px solid var(--border);color:var(--text);
padding:8px 16px;border-radius:8px;font-family:inherit;font-size:14px;outline:none;min-width:200px}
.company-selector select:focus{border-color:var(--accent)}

.upload-zone{border:2px dashed var(--border);border-radius:var(--radius);padding:60px 40px;
text-align:center;transition:all 0.3s;cursor:pointer;background:var(--surface)}
.upload-zone:hover,.upload-zone.dragover{border-color:var(--accent);background:rgba(108,92,231,0.05)}
.upload-icon{width:64px;height:64px;margin:0 auto 20px;background:linear-gradient(135deg,var(--accent),var(--green));
border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:28px}
.upload-title{font-size:20px;font-weight:600;margin-bottom:8px}
.upload-sub{color:var(--text2);font-size:14px}
.upload-input{display:none}

.processing{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(11,15,26,0.85);
backdrop-filter:blur(10px);display:flex;align-items:center;justify-content:center;z-index:200;opacity:0;pointer-events:none;transition:opacity 0.3s}
.processing.active{opacity:1;pointer-events:all}
.processing-card{background:var(--surface);border-radius:var(--radius);padding:48px;text-align:center;border:1px solid var(--border);max-width:400px}
.spinner{width:48px;height:48px;border:3px solid var(--border);border-top-color:var(--accent);border-radius:50%;animation:spin 0.8s linear infinite;margin:0 auto 20px}
@keyframes spin{to{transform:rotate(360deg)}}

.toast{position:fixed;bottom:28px;right:28px;z-index:300;padding:16px 24px;border-radius:var(--radius-sm);
font-weight:500;font-size:14px;transform:translateY(100px);opacity:0;transition:all 0.3s}
.toast.show{transform:translateY(0);opacity:1}
.toast.success{background:var(--green);color:#000}.toast.error{background:var(--red);color:#fff}

.stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:16px;margin-bottom:28px}
.stat-card{background:var(--surface);border-radius:var(--radius);padding:24px;border:1px solid var(--border)}
.stat-label{font-size:13px;color:var(--text2);margin-bottom:8px;text-transform:uppercase;letter-spacing:0.5px}
.stat-value{font-size:32px;font-weight:700;font-family:'JetBrains Mono',monospace}
.stat-value.green{color:var(--green)}

.cat-section{background:var(--surface);border-radius:var(--radius);padding:24px;border:1px solid var(--border);margin-bottom:28px}
.cat-section h3{font-size:16px;margin-bottom:20px;font-weight:600}
.cat-row{display:flex;align-items:center;margin-bottom:14px;gap:12px}
.cat-name{width:160px;font-size:13px;color:var(--text2);flex-shrink:0}
.cat-bar-bg{flex:1;height:28px;background:var(--surface2);border-radius:6px;overflow:hidden}
.cat-bar{height:100%;border-radius:6px;transition:width 0.8s ease;min-width:2px}
.cat-amount{width:100px;text-align:right;font-family:'JetBrains Mono',monospace;font-size:13px;font-weight:500;flex-shrink:0}

.table-wrap{background:var(--surface);border-radius:var(--radius);border:1px solid var(--border);overflow:hidden}
.table-header{display:flex;justify-content:space-between;align-items:center;padding:20px 24px;border-bottom:1px solid var(--border)}
.table-header h3{font-size:16px;font-weight:600}
table{width:100%;border-collapse:collapse}
th{text-align:left;padding:14px 20px;font-size:12px;text-transform:uppercase;letter-spacing:0.5px;
color:var(--text2);font-weight:600;border-bottom:1px solid var(--border);background:var(--surface2)}
td{padding:16px 20px;font-size:14px;border-bottom:1px solid var(--border)}
tr:hover td{background:rgba(108,92,231,0.03)}
.cat-badge{display:inline-block;padding:4px 12px;border-radius:20px;font-size:12px;font-weight:500;
background:rgba(108,92,231,0.15);color:var(--accent2)}
.amount{font-family:'JetBrains Mono',monospace;font-weight:500}
.delete-btn{background:none;border:none;color:var(--text2);cursor:pointer;font-size:16px;
padding:4px 8px;border-radius:6px;transition:all 0.2s}
.delete-btn:hover{color:var(--red);background:rgba(255,107,107,0.1)}
.expense-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:16px;margin-bottom:8px;cursor:pointer;transition:border-color 0.2s}
.expense-card:hover{border-color:var(--accent)}

.empty-state{text-align:center;padding:60px 20px;color:var(--text2)}
.empty-state .icon{font-size:48px;margin-bottom:16px}

.team-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:24px;margin-bottom:16px}
.team-card h3{font-size:16px;font-weight:600;margin-bottom:20px}
.member-row{display:flex;justify-content:space-between;align-items:center;padding:12px 0;border-bottom:1px solid var(--border)}
.member-row:last-child{border-bottom:none}
.member-info{display:flex;flex-direction:column;gap:2px}
.member-name{font-weight:600;font-size:14px}.member-email{font-size:12px;color:var(--text2)}
.role-badge{padding:3px 10px;border-radius:12px;font-size:11px;font-weight:600}
.role-super{background:rgba(253,203,110,0.15);color:var(--orange)}
.role-admin{background:rgba(108,92,231,0.15);color:var(--accent2)}
.role-member{background:rgba(0,210,160,0.15);color:var(--green)}
.invite-code{font-family:'JetBrains Mono',monospace;background:var(--bg);padding:8px 16px;border-radius:8px;font-size:14px;display:inline-block;margin:4px 0}

/* Company cards for super admin */
.company-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);
padding:24px;margin-bottom:16px;display:flex;justify-content:space-between;align-items:center}
.company-info h4{font-size:16px;margin-bottom:4px}.company-stats{font-size:13px;color:var(--text2)}
.company-total{font-family:'JetBrains Mono',monospace;font-size:18px;font-weight:600;color:var(--green)}

.section{display:none}.section.active{display:block}
.modal-overlay{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(11,15,26,0.85);
backdrop-filter:blur(10px);display:flex;align-items:center;justify-content:center;z-index:200;display:none}
.modal-overlay.active{display:flex}
.modal{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:32px;width:100%;max-width:420px;margin:20px}
.modal h3{margin-bottom:20px}
.modal label{font-size:13px;color:var(--text2);display:block;margin-bottom:6px;margin-top:16px}
.modal input,.modal select{width:100%;padding:12px 16px;background:var(--bg);border:1px solid var(--border);
border-radius:10px;color:var(--text);font-family:inherit;font-size:14px;outline:none}

@media(max-width:768px){
.topbar{padding:16px 20px;flex-wrap:wrap;gap:12px}.main{padding:20px}
.nav{padding:0 12px;overflow-x:auto}.nav-tab{padding:14px 16px;font-size:13px;white-space:nowrap}
.upload-zone{padding:40px 20px}.stats-grid{grid-template-columns:repeat(2,1fr);gap:12px}
.stat-value{font-size:24px}.table-wrap{overflow-x:auto}table{min-width:600px}
.cat-name{width:120px}.user-badge{display:none}.company-card{flex-direction:column;gap:12px;align-items:flex-start}}
</style></head><body>

<div class="topbar">
<div class="logo">Expense<span>Snap</span></div>
<div class="topbar-right">
<span class="company-badge">{{ company_name }}</span>
<div class="user-badge">👤 <strong>{{ user_name }}</strong></div>
<button class="btn btn-ghost btn-sm" onclick="exportExcel()">📥 Export</button>
<button class="btn btn-ghost btn-sm" onclick="handleLogout()">Logout</button>
</div>
</div>

<nav class="nav">
<button class="nav-tab active" data-tab="upload">Upload</button>
<button class="nav-tab" data-tab="dashboard">Dashboard</button>
<button class="nav-tab" data-tab="expenses">All Expenses</button>
<button class="nav-tab" data-tab="team">Team</button>
<button class="nav-tab" data-tab="companies" id="companiesTab" style="display:none">Companies</button>
</nav>

<div class="main">

<!-- Company selector for super admin -->
<div class="company-selector" id="companySelector" style="display:none">
<label>👁 Viewing:</label>
<select id="companyFilter" onchange="onCompanyFilterChange()">
<option value="">All Companies</option>
</select>
</div>

<div id="upload" class="section active">
<div class="upload-zone" id="dropZone">
<div class="upload-icon">📸</div>
<div class="upload-title">Drop receipt here or tap to upload</div>
<div class="upload-sub">Supports JPG, PNG, WebP, HEIC, PDF • Phone camera or album</div>
<input type="file" class="upload-input" id="fileInput" accept="image/*,.pdf" multiple>
</div>
<div id="recentUploads" style="margin-top:28px;"></div>
</div>

<div id="dashboard" class="section">
<div class="stats-grid" id="statsGrid"></div>
<div class="cat-section" id="catSection"><h3>Spending by Category</h3><div id="catBars"></div></div>
</div>

<div id="expenses" class="section">
<div class="table-wrap">
<div class="table-header"><h3>All Expenses</h3>
<button class="btn btn-ghost btn-sm" onclick="exportExcel()">📥 Export</button></div>
<div id="expenseTable"></div>
</div>
</div>

<div id="team" class="section">
<div class="team-card"><h3>Invite Team Member</h3>
<p style="color:var(--text2);font-size:14px;margin-bottom:16px;">Generate a one-time invite code to share with a new team member.</p>
<button class="btn btn-primary" onclick="generateInvite()">🔑 Generate Invite Code</button>
<div id="inviteResult" style="margin-top:16px;"></div></div>
<div class="team-card"><h3>Team Members</h3><div id="teamList"></div></div>
<div class="team-card"><h3>Unused Invite Codes</h3><div id="pendingInvites"></div></div>
<div class="team-card" id="companySettingsCard">
<h3>⚙️ Company Settings</h3>
<p style="color:var(--text2);font-size:14px;margin-bottom:16px;">Set your company's home currency. All receipts will be converted to this currency automatically.</p>
<div style="display:flex;gap:12px;align-items:center;flex-wrap:wrap;">
<label style="color:var(--text2);font-size:14px;">Home Currency:</label>
<select id="myCompanyCurrency" style="padding:12px 16px;background:var(--bg);border:1px solid var(--border);border-radius:10px;color:var(--text);font-family:inherit;font-size:14px;outline:none">
<option value="USD">USD $</option><option value="EUR">EUR €</option><option value="GBP">GBP £</option>
<option value="INR">INR ₹</option><option value="CAD">CAD $</option><option value="AUD">AUD $</option>
<option value="SGD">SGD $</option><option value="AED">AED</option><option value="JPY">JPY ¥</option>
<option value="CHF">CHF</option><option value="CNY">CNY ¥</option><option value="MXN">MXN $</option>
</select>
<button class="btn btn-primary" onclick="saveCompanyCurrency()">💾 Save</button>
<button class="btn btn-ghost" onclick="recalculateMyExpenses()">🔄 Recalculate All</button>
</div>
<div id="settingsResult" style="margin-top:12px;"></div>
</div>
</div>

<div id="companies" class="section">
<div class="team-card">
<h3>Create New Company</h3>
<p style="color:var(--text2);font-size:14px;margin-bottom:16px;">Add a new client company. An admin invite code will be auto-generated.</p>
<div style="display:flex;gap:12px;align-items:center;flex-wrap:wrap;">
<input type="text" id="newCompanyName" placeholder="Company name" style="flex:1;min-width:200px;padding:12px 16px;
background:var(--bg);border:1px solid var(--border);border-radius:10px;color:var(--text);font-family:inherit;font-size:14px;outline:none">
<select id="newCompanyCurrency" style="padding:12px 16px;background:var(--bg);border:1px solid var(--border);border-radius:10px;color:var(--text);font-family:inherit;font-size:14px;outline:none">
<option value="USD">USD $</option><option value="EUR">EUR €</option><option value="GBP">GBP £</option>
<option value="INR">INR ₹</option><option value="CAD">CAD $</option><option value="AUD">AUD $</option>
<option value="SGD">SGD $</option><option value="AED">AED</option><option value="JPY">JPY ¥</option>
<option value="CHF">CHF</option><option value="CNY">CNY ¥</option><option value="MXN">MXN $</option>
</select>
<button class="btn btn-primary" onclick="createCompany()">+ Create Company</button>
</div>
<div id="companyResult" style="margin-top:16px;"></div>
</div>
<div id="companyList"></div>
</div>

</div>

<div class="processing" id="processing">
<div class="processing-card"><div class="spinner"></div>
<div style="font-size:18px;font-weight:600;margin-bottom:8px;">Scanning Receipt...</div>
<div style="color:var(--text2);font-size:14px;">Claude is reading your receipt</div></div>
</div>

<div class="toast" id="toast"></div>

<script>
const USER_ROLE = '{{ user_role }}';
const userRole = USER_ROLE;
const isSuperAdmin = USER_ROLE === 'super_admin';
const myCompanyId = '{{ company_id }}' === 'None' ? '' : '{{ company_id }}';
let selectedCompany = '';

// Show super admin UI
if (isSuperAdmin) {
  document.getElementById('companiesTab').style.display = 'block';
  document.getElementById('companySelector').style.display = 'flex';
  loadCompanyFilter();
}

function onCompanyFilterChange() {
  selectedCompany = document.getElementById('companyFilter').value;
  // Reload current tab data
  const activeTab = document.querySelector('.nav-tab.active').dataset.tab;
  if (activeTab === 'dashboard') loadDashboard();
  if (activeTab === 'expenses') loadExpenses();
}

async function loadCompanyFilter() {
  const res = await fetch('/api/companies');
  const companies = await res.json();
  const sel = document.getElementById('companyFilter');
  sel.innerHTML = '<option value="">All Companies</option>';
  companies.forEach(c => { sel.innerHTML += `<option value="${c.id}">${c.name}</option>`; });
}

function apiUrl(base) {
  if (isSuperAdmin && selectedCompany) return base + (base.includes('?') ? '&' : '?') + 'company_id=' + selectedCompany;
  return base;
}

// Navigation
document.querySelectorAll('.nav-tab').forEach(tab => {
  tab.addEventListener('click', () => {
    document.querySelectorAll('.nav-tab').forEach(t => t.classList.remove('active'));
    document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
    tab.classList.add('active');
    document.getElementById(tab.dataset.tab).classList.add('active');
    if (tab.dataset.tab === 'dashboard') loadDashboard();
    if (tab.dataset.tab === 'expenses') loadExpenses();
    if (tab.dataset.tab === 'team') { loadTeam(); loadCompanySettings(); }
    if (tab.dataset.tab === 'companies') loadCompanies();
  });
});

// Upload
const dropZone = document.getElementById('dropZone');
const fileInput = document.getElementById('fileInput');
dropZone.addEventListener('click', () => fileInput.click());
dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('dragover'); });
dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
dropZone.addEventListener('drop', e => { e.preventDefault(); dropZone.classList.remove('dragover'); handleFiles(e.dataTransfer.files); });
fileInput.addEventListener('change', e => handleFiles(e.target.files));

async function handleFiles(files) { for (const file of files) await uploadFile(file); fileInput.value = ''; }

async function uploadFile(file) {
  document.getElementById('processing').classList.add('active');
  const fd = new FormData(); fd.append('receipt', file);
  try {
    const res = await fetch('/api/upload', {method:'POST', body:fd});
    const data = await res.json();
    if (data.success) { showToast(`✓ ${data.expense.vendor} — ${data.expense.currency} ${data.expense.total}`, 'success'); showRecentUpload(data.expense); }
    else { if (res.status===401) { window.location.href='/login'; return; } showToast('Failed: '+(data.error||'Unknown error'),'error'); }
  } catch(err) { showToast('Upload failed: '+err.message, 'error'); }
  document.getElementById('processing').classList.remove('active');
}

function showRecentUpload(exp) {
  const c = document.getElementById('recentUploads');
  const d = document.createElement('div');
  d.style.cssText = 'background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:20px;margin-bottom:12px;display:flex;justify-content:space-between;align-items:center;';
  d.innerHTML = `<div><div style="font-weight:600;margin-bottom:4px;">${exp.vendor||'Unknown'}</div>
  <div style="font-size:13px;color:var(--text2);">${exp.date} · ${exp.category} · by ${exp.uploaded_by||''}</div></div>
  <div style="font-family:'JetBrains Mono',monospace;font-size:18px;font-weight:600;color:var(--green);">${exp.currency} ${Number(exp.total).toFixed(2)}</div>`;
  c.prepend(d);
}

// Dashboard
async function loadDashboard() {
  try {
    const res = await fetch(apiUrl('/api/dashboard'));
    if (res.status===401) { window.location.href='/login'; return; }
    const data = await res.json();
    document.getElementById('statsGrid').innerHTML = `
      <div class="stat-card"><div class="stat-label">Bill Total</div><div class="stat-value green">${data.total.toFixed(2)}</div><div style="color:var(--text2);font-size:12px;margin-top:4px">Original currencies</div></div>
      <div class="stat-card"><div class="stat-label">Home Currency</div><div class="stat-value" style="color:var(--accent2)">${data.home_currency} ${data.total_home.toFixed(2)}</div><div style="color:var(--text2);font-size:12px;margin-top:4px">Converted total</div></div>
      <div class="stat-card"><div class="stat-label">USD Total</div><div class="stat-value" style="color:#FFD93D">USD ${data.total_usd.toFixed(2)}</div><div style="color:var(--text2);font-size:12px;margin-top:4px">Global currency</div></div>
      <div class="stat-card"><div class="stat-label">Receipts</div><div class="stat-value">${data.count}</div><div style="color:var(--text2);font-size:12px;margin-top:4px">${Object.keys(data.by_category).length} categories</div></div>`;
    const cats = Object.entries(data.by_category).sort((a,b)=>b[1]-a[1]);
    const maxVal = cats.length ? cats[0][1] : 1;
    const colors = ['#6C5CE7','#00D2A0','#FDCB6E','#74B9FF','#FF6B6B','#A29BFE','#FD79A8','#55E6C1'];
    document.getElementById('catBars').innerHTML = cats.length ? cats.map(([cat,amt],i) => `
      <div class="cat-row"><div class="cat-name">${cat}</div>
      <div class="cat-bar-bg"><div class="cat-bar" style="width:${amt/maxVal*100}%;background:${colors[i%colors.length]};"></div></div>
      <div class="cat-amount">${data.home_currency} ${amt.toFixed(2)}</div></div>`).join('') : '<div class="empty-state"><p>No expenses yet</p></div>';
  } catch(e) { console.error(e); }
}

// Expenses
async function loadExpenses() {
  try {
    const res = await fetch(apiUrl('/api/expenses'));
    if (res.status===401) { window.location.href='/login'; return; }
    const expenses = await res.json();
    if (!expenses.length) { document.getElementById('expenseTable').innerHTML = '<div class="empty-state"><div class="icon">🧾</div><p>No expenses yet</p></div>'; return; }
    const showCompany = isSuperAdmin && !selectedCompany;
    document.getElementById('expenseTable').innerHTML = expenses.map(e => `
      <div class="expense-card" onclick="toggleExpenseDetail(this)">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:12px;">
          <div style="flex:1;">
            <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;">
              <strong>${e.vendor}</strong>
              <span class="cat-badge">${e.category}</span>
              ${showCompany?`<span style="font-size:11px;color:var(--accent2);background:rgba(99,102,241,0.15);padding:2px 8px;border-radius:8px;">${e.company_name||''}</span>`:''}
            </div>
            <div style="font-size:12px;color:var(--text2);margin-top:4px;">${e.date} · ${e.location||''} ${e.uploaded_by?'· '+e.uploaded_by:''}</div>
          </div>
          <div style="text-align:right;">
            <div class="amount" style="font-size:18px;">${e.currency} ${Number(e.total).toFixed(2)}</div>
            ${e.total_home?`<div style="font-size:11px;color:var(--text2);">Home: ${Number(e.total_home).toFixed(2)} · USD: ${Number(e.total_usd).toFixed(2)}</div>`:''}
          </div>
        </div>
        <div class="expense-detail" style="display:none;margin-top:12px;padding-top:12px;border-top:1px solid var(--border);">
          <div style="display:flex;gap:24px;flex-wrap:wrap;margin-bottom:8px;">
            <div><span style="color:var(--text2);font-size:12px;">Subtotal</span><br><strong>${e.currency} ${Number(e.subtotal||0).toFixed(2)}</strong></div>
            <div><span style="color:var(--text2);font-size:12px;">Tax</span><br><strong style="color:#f59e0b;">${e.currency} ${Number(e.tax||0).toFixed(2)}</strong></div>
            <div><span style="color:var(--text2);font-size:12px;">Tip</span><br><strong>${e.currency} ${Number(e.tip||0).toFixed(2)}</strong></div>
            <div><span style="color:var(--text2);font-size:12px;">Payment</span><br><strong>${e.payment_method||'N/A'}</strong></div>
          </div>
          ${e.items?`<div style="margin-top:8px;"><span style="color:var(--text2);font-size:12px;">Items</span><br><div style="font-size:13px;color:var(--text);margin-top:4px;line-height:1.6;">${e.items}</div></div>`:''}
          <div style="margin-top:8px;text-align:right;"><button class="delete-btn" onclick="event.stopPropagation();deleteExpense('${e.id}')" style="font-size:12px;color:#ef4444;background:rgba(239,68,68,0.1);border:none;padding:4px 12px;border-radius:6px;cursor:pointer;">🗑 Delete</button></div>
        </div>
      </div>`).join('');
  } catch(e) { console.error(e); }
}

function toggleExpenseDetail(card) {
  const detail = card.querySelector('.expense-detail');
  if (detail) detail.style.display = detail.style.display === 'none' ? 'block' : 'none';
}

async function deleteExpense(id) {
  if (!confirm('Delete this expense?')) return;
  await fetch(`/api/expenses/${id}`,{method:'DELETE'});
  loadExpenses(); showToast('Expense deleted','success');
}

// Team
async function loadTeam() {
  try {
    let url = '/api/team';
    if (isSuperAdmin && selectedCompany) url += '?company_id='+selectedCompany;
    const res = await fetch(url);
    if (res.status===401||res.status===403) return;
    const data = await res.json();
    document.getElementById('teamList').innerHTML = data.users.map(u => `
      <div class="member-row"><div class="member-info"><div class="member-name">${u.name}</div>
      <div class="member-email">${u.email}</div></div>
      <div style="display:flex;align-items:center;gap:8px;">
      <span class="role-badge ${u.role==='super_admin'?'role-super':u.role==='company_admin'?'role-admin':'role-member'}">${u.role.replace('_',' ')}</span>
      ${u.role!=='super_admin'?`<button class="btn btn-ghost btn-sm" onclick="resetPassword('${u.id}','${u.name}')">🔑 Reset</button><button class="btn btn-danger btn-sm" onclick="removeMember('${u.id}')">Remove</button>`:''}</div></div>`).join('')
      || '<p style="color:var(--text2);font-size:14px;">No team members yet</p>';
    document.getElementById('pendingInvites').innerHTML = data.pending_invites.length ?
      data.pending_invites.map(i => `<div class="invite-code">${i.code} <span style="color:var(--text2);font-size:11px">(${i.role||'member'})</span></div>`).join(' ') :
      '<p style="color:var(--text2);font-size:14px;">No pending invites</p>';
  } catch(e) { console.error(e); }
}

async function generateInvite() {
  let body = {};
  if (isSuperAdmin) {
    const cid = selectedCompany || prompt('Enter company ID (check Companies tab):');
    if (!cid) return;
    const role = confirm('Make them a Company Admin? (OK=Admin, Cancel=Member)') ? 'company_admin' : 'member';
    body = {company_id: cid, role: role};
  }
  const res = await fetch('/api/invite',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
  const data = await res.json();
  if (data.success) {
    document.getElementById('inviteResult').innerHTML = `
      <div style="background:rgba(0,210,160,0.1);padding:16px;border-radius:10px;">
      <div style="font-size:13px;color:var(--green);margin-bottom:8px;">✓ Invite code created!</div>
      <div class="invite-code" style="font-size:18px;">${data.code}</div>
      <div style="font-size:12px;color:var(--text2);margin-top:8px;">Share this with your team member. One-time use only.</div></div>`;
    loadTeam();
  } else { showToast(data.error,'error'); }
}

async function removeMember(id) {
  if (!confirm('Remove this team member?')) return;
  await fetch(`/api/team/${id}`,{method:'DELETE'});
  loadTeam(); showToast('Member removed','success');
}

async function resetPassword(id, name) {
  const pwd = prompt(`Enter new password for ${name} (min 6 characters):`);
  if (!pwd || pwd.length < 6) { if(pwd) showToast('Password must be at least 6 characters','error'); return; }
  const res = await fetch(`/api/team/${id}/reset-password`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({password:pwd})});
  const data = await res.json();
  if (data.success) showToast(`Password reset for ${name}`,'success');
  else showToast(data.error,'error');
}

// Companies (Super Admin)
async function loadCompanies() {
  const res = await fetch('/api/companies');
  const companies = await res.json();
  document.getElementById('companyList').innerHTML = companies.map(c => `
    <div class="company-card">
    <div><div class="company-info"><h4>${c.name}</h4></div>
    <div class="company-stats">${c.user_count} users · ${c.expense_count} receipts · ${c.home_currency||'USD'}</div></div>
    <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;">
    <div class="company-total">${c.home_currency||'USD'} ${c.total_spent.toFixed(2)}</div>
    <button class="btn btn-ghost btn-sm" onclick="editCompanyCurrency('${c.id}','${c.name}','${c.home_currency||'USD'}')">💱 Currency</button>
    <button class="btn btn-ghost btn-sm" onclick="recalculateCompany('${c.id}','${c.name}')">🔄 Recalc</button>
    <button class="btn btn-danger btn-sm" onclick="deleteCompany('${c.id}','${c.name}')">Delete</button></div>
    </div>`).join('') || '<div class="empty-state"><p>No companies yet. Create one above!</p></div>';
}

async function createCompany() {
  const name = document.getElementById('newCompanyName').value.trim();
  const home_currency = document.getElementById('newCompanyCurrency').value;
  if (!name) { showToast('Enter a company name','error'); return; }
  const res = await fetch('/api/companies',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name, home_currency})});
  const data = await res.json();
  if (data.success) {
    document.getElementById('companyResult').innerHTML = `
      <div style="background:rgba(0,210,160,0.1);padding:16px;border-radius:10px;">
      <div style="font-size:13px;color:var(--green);margin-bottom:8px;">✓ Company "${name}" created! (${home_currency})</div>
      <div style="margin-bottom:4px;">Admin invite code:</div>
      <div class="invite-code" style="font-size:18px;">${data.admin_invite_code}</div>
      <div style="font-size:12px;color:var(--text2);margin-top:8px;">Send this to the company's main person. They'll register and become the admin.</div></div>`;
    document.getElementById('newCompanyName').value = '';
    loadCompanies(); loadCompanyFilter();
  } else { showToast(data.error,'error'); }
}

async function deleteCompany(id,name) {
  if (!confirm(`Delete "${name}" and ALL its data? This cannot be undone!`)) return;
  await fetch(`/api/companies/${id}`,{method:'DELETE'});
  loadCompanies(); loadCompanyFilter(); showToast(`${name} deleted`,'success');
}

async function editCompanyCurrency(id, name, currentCurrency) {
  const currencies = ['USD','EUR','GBP','INR','CAD','AUD','SGD','AED','JPY','CHF','CNY','MXN'];
  const curr = prompt(`Change home currency for "${name}"\nCurrent: ${currentCurrency}\n\nEnter new currency code (${currencies.join(', ')}):`);
  if (!curr) return;
  if (!currencies.includes(curr.toUpperCase())) { showToast('Invalid currency code','error'); return; }
  const res = await fetch(`/api/companies/${id}`,{method:'PUT',headers:{'Content-Type':'application/json'},body:JSON.stringify({home_currency:curr.toUpperCase()})});
  const data = await res.json();
  if (data.success) { showToast(`${name} currency changed to ${curr.toUpperCase()}`,'success'); loadCompanies(); }
  else showToast(data.error,'error');
}

async function recalculateCompany(id, name) {
  if (!confirm(`Recalculate all converted amounts for "${name}" using current exchange rates?`)) return;
  const res = await fetch(`/api/companies/${id}/recalculate`,{method:'POST'});
  const data = await res.json();
  if (data.success) { showToast(`${data.updated} receipts recalculated to ${data.home_currency}`,'success'); loadCompanies(); loadDashboard(); }
  else showToast(data.error,'error');
}

// Company Settings (for Company Admins)
async function loadCompanySettings() {
  const card = document.getElementById('companySettingsCard');
  if (!card) return;
  if (userRole === 'member') { card.style.display = 'none'; return; }
  try {
    const cid = selectedCompany || myCompanyId;
    if (isSuperAdmin && !selectedCompany) { card.style.display = 'none'; return; }
    const res = await fetch('/api/my-company');
    if (!res.ok) { card.style.display = 'none'; return; }
    const comp = await res.json();
    const sel = document.getElementById('myCompanyCurrency');
    if (sel && comp.home_currency) sel.value = comp.home_currency;
  } catch(e) { console.error(e); }
}

async function saveCompanyCurrency() {
  const curr = document.getElementById('myCompanyCurrency').value;
  const cid = selectedCompany || myCompanyId;
  if (!cid || cid === 'None') { showToast('No company selected','error'); return; }
  const res = await fetch(`/api/companies/${cid}`,{method:'PUT',headers:{'Content-Type':'application/json'},body:JSON.stringify({home_currency:curr})});
  const data = await res.json();
  if (data.success) {
    document.getElementById('settingsResult').innerHTML = `<div style="color:var(--green);font-size:13px;">✓ Home currency set to ${curr}</div>`;
    showToast(`Home currency changed to ${curr}`,'success');
  } else showToast(data.error,'error');
}

async function recalculateMyExpenses() {
  const cid = selectedCompany || myCompanyId;
  if (!cid || cid === 'None') { showToast('No company selected','error'); return; }
  if (!confirm('Recalculate all receipts with current exchange rates?')) return;
  const res = await fetch(`/api/companies/${cid}/recalculate`,{method:'POST'});
  const data = await res.json();
  if (data.success) {
    document.getElementById('settingsResult').innerHTML = `<div style="color:var(--green);font-size:13px;">✓ ${data.updated} receipts recalculated to ${data.home_currency}</div>`;
    showToast(`${data.updated} receipts recalculated`,'success'); loadDashboard();
  } else showToast(data.error,'error');
}

// Utilities
async function handleLogout() { await fetch('/api/logout',{method:'POST'}); window.location.href='/login'; }
function exportExcel() { window.location.href = apiUrl('/api/export'); }
function showToast(msg,type='success') {
  const t = document.getElementById('toast'); t.textContent=msg; t.className=`toast ${type} show`;
  setTimeout(()=>t.classList.remove('show'),3500);
}
</script></body></html>"""

init_db()

# --- External API for FinanceSnap ---
@app.route('/api/expenses/external')
def api_expenses_external():
    api_key = request.headers.get('X-API-Key', '')
    if not api_key:
        return jsonify({'error': 'API key required'}), 401
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE email=%s", (api_key,))
    user = cur.fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'Invalid API key'}), 401

    company_id = request.args.get('company_id', '')

    if user['role'] == 'super_admin':
        if company_id:
            cur.execute("SELECT e.*, c.name as company_name, c.home_currency FROM expenses e LEFT JOIN companies c ON e.company_id=c.id WHERE e.company_id=%s ORDER BY e.date DESC", (company_id,))
        else:
            cur.execute("SELECT e.*, c.name as company_name, c.home_currency FROM expenses e LEFT JOIN companies c ON e.company_id=c.id ORDER BY e.date DESC")
    else:
        cid = company_id or user['company_id']
        if cid:
            cur.execute("SELECT e.*, c.name as company_name, c.home_currency FROM expenses e LEFT JOIN companies c ON e.company_id=c.id WHERE e.company_id=%s ORDER BY e.date DESC", (cid,))
        else:
            cur.execute("SELECT e.*, c.name as company_name, c.home_currency FROM expenses e LEFT JOIN companies c ON e.company_id=c.id ORDER BY e.date DESC LIMIT 100")
    rows = cur.fetchall()
    conn.close()
    expenses = []
    for r in rows:
        d = dict(r)
        for k, v in d.items():
            if hasattr(v, 'isoformat'):
                d[k] = v.isoformat()
        expenses.append(d)
    return jsonify({'expenses': expenses, 'count': len(expenses)})

@app.route('/api/companies/external')
def api_companies_external():
    api_key = request.headers.get('X-API-Key', '')
    if not api_key:
        return jsonify({'error': 'API key required'}), 401
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE email=%s", (api_key,))
    user = cur.fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'Invalid API key'}), 401
    if user['role'] != 'super_admin':
        conn.close()
        return jsonify({'error': 'Admin only'}), 403

    cur.execute("""SELECT c.*,
        COUNT(e.id) as receipt_count,
        COALESCE(SUM(e.total), 0) as total_expenses,
        (SELECT COUNT(*) FROM users u WHERE u.company_id = c.id) as user_count
        FROM companies c LEFT JOIN expenses e ON c.id = e.company_id
        GROUP BY c.id ORDER BY c.name""")
    companies = cur.fetchall()
    conn.close()
    result = []
    for c in companies:
        d = dict(c)
        for k, v in d.items():
            if hasattr(v, 'isoformat'):
                d[k] = v.isoformat()
        result.append(d)
    return jsonify({'companies': result, 'count': len(result)})

if __name__ == '__main__':
    print("\n" + "="*50)
    print("  🧾 ExpenseSnap is running!")
    print("="*50)
    print(f"  Open: http://localhost:5000")
    print(f"  Press Ctrl+C to stop")
    print("="*50 + "\n")
    app.run(host='0.0.0.0', port=5000, debug=True)
